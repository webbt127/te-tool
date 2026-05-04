import argparse
import hashlib
import os
import re
import sqlite3
from datetime import date, datetime

import openpyxl
from tqdm import tqdm

DB_NAME = "timesheet.db"
__version__ = "0.1-alpha"


# -----------------------------
# General helpers
# -----------------------------

def file_hash(path: str) -> str:
    """Return an MD5 hash for a file so changed spreadsheets can be detected."""
    h = hashlib.md5()
    with open(path, "rb") as f:
        while chunk := f.read(8192):
            h.update(chunk)
    return h.hexdigest()


def get_conn():
    return sqlite3.connect(DB_NAME)


def safe_float(value) -> float:
    """Convert numbers, currency strings, and blanks into a float."""
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    cleaned = str(value).replace("$", "").replace(",", "").strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def is_true(value) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in ("true", "yes", "y", "1")


def is_weekday_only(value) -> bool:
    return str(value or "").strip().lower() in (
        "monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday",
        "mon", "tue", "wed", "thu", "fri", "sat", "sun",
    )


def parse_date_value(value):
    """
    Return an ISO date string (YYYY-MM-DD) if value looks like a date.
    Return None for blanks, weekdays, names, and other non-date values.
    """
    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if value is None or value == "":
        return None

    if is_weekday_only(value):
        return None

    text = str(value).strip()
    for fmt in (
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%m-%d-%Y",
        "%m-%d-%y",
        "%B %d, %Y",
        "%b %d, %Y",
    ):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass

    return None


def normalize_group_date(raw_date, group_rows):
    """
    New sheets usually have the real date on the 3rd row of the 5-row group.
    Older sheets may have only the weekday there, so scan the entire group for
    a real date value.
    """
    parsed = parse_date_value(raw_date)
    if parsed:
        return parsed

    for row in group_rows:
        if not row:
            continue
        for cell in row:
            parsed = parse_date_value(cell)
            if parsed:
                return parsed

    return ""


def expense_total_sql(prefix: str = "e") -> str:
    return f"""
        COALESCE({prefix}.mileage_dollars, 0) +
        COALESCE({prefix}.per_diem_food, 0) +
        COALESCE({prefix}.air_fare, 0) +
        COALESCE({prefix}.hotel, 0) +
        COALESCE({prefix}.parking_tolls, 0) +
        COALESCE({prefix}.rental_car_fuel, 0) +
        COALESCE({prefix}.business_meals, 0) +
        COALESCE({prefix}.other, 0)
    """


def print_table(headers, rows):
    """Print rows as a simple aligned text table."""
    if not rows:
        print("No rows to display.")
        return

    str_rows = [[str(value if value is not None else "") for value in row] for row in rows]
    widths = []

    for index, header in enumerate(headers):
        max_cell_width = max(len(row[index]) for row in str_rows) if str_rows else 0
        widths.append(max(len(header), max_cell_width))

    print(" | ".join(header.ljust(widths[index]) for index, header in enumerate(headers)))
    print("-+-".join("-" * width for width in widths))

    for row in str_rows:
        print(" | ".join(row[index].ljust(widths[index]) for index in range(len(headers))))


# -----------------------------
# Database setup / migration
# -----------------------------

def init_db():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS config (
            key TEXT PRIMARY KEY,
            value TEXT
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS files (
            path TEXT PRIMARY KEY,
            hash TEXT,
            last_processed TEXT
        )
        """
    )

    ensure_time_schema(cur)
    ensure_expense_schema(cur)

    conn.commit()
    conn.close()


def ensure_time_schema(cur):
    """
    Ensure time_entries uses the current schema.

    SQLite does not have a strict DATE storage class. Declaring the column as
    DATE gives it DATE affinity; values are stored as ISO date strings.
    """
    expected_columns = {
        "id",
        "date",
        "job_number",
        "work_code",
        "regular_hours",
        "ot_hours",
        "doubletime_hours",
        "description",
        "overnight",
        "source_file",
    }

    cur.execute("PRAGMA table_info(time_entries)")
    table_info = cur.fetchall()
    existing_columns = {row[1] for row in table_info}
    date_column_type = ""

    for row in table_info:
        if row[1] == "date":
            date_column_type = str(row[2] or "").upper()
            break

    if existing_columns and (
        not expected_columns.issubset(existing_columns)
        or date_column_type != "DATE"
    ):
        cur.execute("DROP TABLE IF EXISTS expense_entries")
        cur.execute("DROP TABLE IF EXISTS time_entries")
        cur.execute("DELETE FROM files")

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS time_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date DATE,
            job_number TEXT,
            work_code TEXT,
            regular_hours REAL,
            ot_hours REAL,
            doubletime_hours REAL,
            description TEXT,
            overnight INTEGER,
            source_file TEXT
        )
        """
    )


def ensure_expense_schema(cur):
    """
    Ensure expense_entries uses the current schema.

    The expense table intentionally does not store date or job_number. Those
    belong to the associated time_entries record via time_entry_id.
    """
    expected_columns = {
        "id",
        "time_entry_id",
        "mileage",
        "mileage_dollars",
        "per_diem_food",
        "air_fare",
        "hotel",
        "parking_tolls",
        "rental_car_fuel",
        "business_meals",
        "other",
        "explanation",
        "source_file",
    }

    cur.execute("PRAGMA table_info(expense_entries)")
    existing_columns = {row[1] for row in cur.fetchall()}

    if existing_columns and not expected_columns.issubset(existing_columns):
        cur.execute("DROP TABLE IF EXISTS expense_entries")
        cur.execute("DELETE FROM files")

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS expense_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            time_entry_id INTEGER,
            mileage REAL,
            mileage_dollars REAL,
            per_diem_food REAL,
            air_fare REAL,
            hotel REAL,
            parking_tolls REAL,
            rental_car_fuel REAL,
            business_meals REAL,
            other REAL,
            explanation TEXT,
            source_file TEXT,
            FOREIGN KEY(time_entry_id) REFERENCES time_entries(id)
        )
        """
    )


# -----------------------------
# Config
# -----------------------------

def set_config(directory: str, user: str = None):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "REPLACE INTO config (key, value) VALUES (?, ?)",
        ("timesheet_dir", directory),
    )
    if user:
        cur.execute(
            "REPLACE INTO config (key, value) VALUES (?, ?)",
            ("user", user),
        )
    conn.commit()
    conn.close()
    print(f"Configured directory: {directory}")


def get_config(key: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT value FROM config WHERE key=?", (key,))
    row = cur.fetchone()
    conn.close()
    return row[0] if row else None


def validate_config(require_dir=True, require_user=True):
    directory = get_config("timesheet_dir")
    user = get_config("user")

    if require_dir and not directory:
        print("Error: timesheet directory is not configured.")
        print("Run: python timesheet_expense_tool.py config <directory> --user <name>")
        return False

    if require_user and not user:
        print("Error: user is not configured.")
        print("Run: python timesheet_expense_tool.py config <directory> --user <name>")
        return False

    return True


# -----------------------------
# Excel parsing
# -----------------------------

def parse_timesheet(file_path: str):
    """
    Parse a workbook assuming a TRUE 1:1 row relationship between the time
    worksheet and the expense worksheet.

    Time worksheet rules:
    - Time data is read from rows 20 through 54.
    - Time rows are grouped in 5-row blocks.
    - The date for each 5-row block comes from the 3rd row of that block.
      Older sheets that have a weekday there are handled by scanning the group
      for a real date.
    - The overnight value comes from column N on the 5th row of that block.
    - Rows with no job number AND no work code are skipped.

    Expense worksheet rules:
    - Expense data is read from rows 9 through 43.
    - Expense rows are also grouped in 5-row blocks.
    - Expense columns are:
        C Miles                mileage
        D Mileage dollars      mileage_dollars
        E Per Diem Food        per_diem_food
        F Air Fare             air_fare
        G Hotel                hotel
        H Parking / Tolls      parking_tolls
        I Rental Car / Fuel    rental_car_fuel
        J Business Meals       business_meals
        K Other Expense        other
        L Explanation          explanation
    """
    time_min_row = 20
    time_max_row = 54
    expense_min_row = 9
    expense_max_row = 43

    group_size = 5
    date_offset_in_group = 2       # 3rd row, zero-based
    overnight_offset_in_group = 4  # 5th row, zero-based

    wb = openpyxl.load_workbook(file_path, data_only=True)

    time_sheet = wb[wb.sheetnames[0]]
    expense_sheet = wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else None

    time_rows = list(
        time_sheet.iter_rows(
            min_row=time_min_row,
            max_row=time_max_row,
            values_only=True,
        )
    )

    expense_rows = (
        list(
            expense_sheet.iter_rows(
                min_row=expense_min_row,
                max_row=expense_max_row,
                values_only=True,
            )
        )
        if expense_sheet
        else []
    )

    time_entries = []
    expense_entries = []

    for group_start in range(0, len(time_rows), group_size):
        group_time_rows = time_rows[group_start:group_start + group_size]
        group_expense_rows = expense_rows[group_start:group_start + group_size]

        first_time_excel_row = time_min_row + group_start
        date_excel_row = first_time_excel_row + date_offset_in_group

        raw_group_date = ""
        if len(group_time_rows) > date_offset_in_group:
            date_row = group_time_rows[date_offset_in_group]
            if date_row:
                raw_group_date = date_row[0]

        group_date = normalize_group_date(raw_group_date, group_time_rows)
        if not group_date:
            tqdm.write(
                f"Skipping rows {first_time_excel_row}-{first_time_excel_row + group_size - 1} "
                f"in {file_path}: no real date found for the 5-row group. "
                f"Expected date around row {date_excel_row}."
            )
            continue

        overnight_value = None
        if len(group_time_rows) > overnight_offset_in_group:
            overnight_row = group_time_rows[overnight_offset_in_group]
            if overnight_row and len(overnight_row) > 13:
                overnight_value = overnight_row[13]  # Column N

        overnight = 1 if is_true(overnight_value) else 0

        for offset, t_row in enumerate(group_time_rows):
            source_row_number = first_time_excel_row + offset
            e_row = group_expense_rows[offset] if offset < len(group_expense_rows) else None

            if not t_row or not any(t_row):
                time_entries.append(None)
                expense_entries.append(None)
                continue

            job_number = t_row[1] if len(t_row) > 1 else None
            work_code = t_row[2] if len(t_row) > 2 else None

            if not job_number and not work_code:
                time_entries.append(None)
                expense_entries.append(None)
                continue

            try:
                regular_hours = float(t_row[3] or 0)
                ot_hours = float(t_row[4] or 0)
                doubletime_hours = float(t_row[5] or 0)
            except (TypeError, ValueError):
                tqdm.write(
                    f"Skipping row {source_row_number} in {file_path}: "
                    f"hour columns are not numeric. Row data: {t_row}"
                )
                time_entries.append(None)
                expense_entries.append(None)
                continue

            time_entries.append({
                "date": group_date,
                "job_number": str(job_number or ""),
                "work_code": str(work_code or ""),
                "regular_hours": regular_hours,
                "ot_hours": ot_hours,
                "doubletime_hours": doubletime_hours,
                "description": str(t_row[6] or ""),
                "overnight": overnight,
            })

            if e_row and any(e_row):
                expense_entries.append({
                    "mileage": safe_float(e_row[2] if len(e_row) > 2 else 0),
                    "mileage_dollars": safe_float(e_row[3] if len(e_row) > 3 else 0),
                    "per_diem_food": safe_float(e_row[4] if len(e_row) > 4 else 0),
                    "air_fare": safe_float(e_row[5] if len(e_row) > 5 else 0),
                    "hotel": safe_float(e_row[6] if len(e_row) > 6 else 0),
                    "parking_tolls": safe_float(e_row[7] if len(e_row) > 7 else 0),
                    "rental_car_fuel": safe_float(e_row[8] if len(e_row) > 8 else 0),
                    "business_meals": safe_float(e_row[9] if len(e_row) > 9 else 0),
                    "other": safe_float(e_row[10] if len(e_row) > 10 else 0),
                    "explanation": str(e_row[11] or "") if len(e_row) > 11 else "",
                })
            else:
                expense_entries.append(None)

    return time_entries, expense_entries


# -----------------------------
# Rebase
# -----------------------------

def find_xlsx_files(directory: str):
    xlsx_files = []
    for root, _, files in os.walk(directory):
        for filename in files:
            if (
                filename.lower().endswith(".xlsx")
                and not filename.startswith("~$")
                and "XXXX" not in filename.upper()
            ):
                xlsx_files.append(os.path.join(root, filename))
    return xlsx_files


def clear_imported_data():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM expense_entries")
    cur.execute("DELETE FROM time_entries")
    cur.execute("DELETE FROM files")
    conn.commit()
    conn.close()


def rebase():
    directory = get_config("timesheet_dir")
    if not directory:
        print("No directory configured. Run: python timesheet_expense_tool.py config <directory>")
        return

    if not os.path.isdir(directory):
        print(f"Configured directory does not exist: {directory}")
        return

    all_files = find_xlsx_files(directory)
    total_files = len(all_files)

    processed = 0
    skipped = 0
    errors = 0
    total_time_rows = 0
    total_expense_rows = 0

    print(f"Found {total_files} Excel file(s). Starting rebase...")

    conn = get_conn()
    cur = conn.cursor()

    with tqdm(total=total_files, desc="Rebasing files", unit="file") as pbar:
        for path in all_files:
            try:
                current_hash = file_hash(path)

                cur.execute("SELECT hash FROM files WHERE path=?", (path,))
                row = cur.fetchone()

                if row and row[0] == current_hash:
                    skipped += 1
                    pbar.set_postfix_str("skipped")
                    pbar.update(1)
                    continue

                pbar.set_postfix_str("processing")

                cur.execute("DELETE FROM expense_entries WHERE source_file=?", (path,))
                cur.execute("DELETE FROM time_entries WHERE source_file=?", (path,))

                time_entries, expense_entries = parse_timesheet(path)

                inserted_time = 0
                inserted_expense = 0

                for time_entry, expense_entry in zip(time_entries, expense_entries):
                    if not time_entry:
                        continue

                    cur.execute(
                        """
                        INSERT INTO time_entries (
                            date,
                            job_number,
                            work_code,
                            regular_hours,
                            ot_hours,
                            doubletime_hours,
                            description,
                            overnight,
                            source_file
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            time_entry["date"],
                            time_entry["job_number"],
                            time_entry["work_code"],
                            time_entry["regular_hours"],
                            time_entry["ot_hours"],
                            time_entry["doubletime_hours"],
                            time_entry["description"],
                            time_entry["overnight"],
                            path,
                        ),
                    )
                    time_entry_id = cur.lastrowid
                    inserted_time += 1

                    if expense_entry:
                        cur.execute(
                            """
                            INSERT INTO expense_entries (
                                time_entry_id,
                                mileage,
                                mileage_dollars,
                                per_diem_food,
                                air_fare,
                                hotel,
                                parking_tolls,
                                rental_car_fuel,
                                business_meals,
                                other,
                                explanation,
                                source_file
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                time_entry_id,
                                expense_entry["mileage"],
                                expense_entry["mileage_dollars"],
                                expense_entry["per_diem_food"],
                                expense_entry["air_fare"],
                                expense_entry["hotel"],
                                expense_entry["parking_tolls"],
                                expense_entry["rental_car_fuel"],
                                expense_entry["business_meals"],
                                expense_entry["other"],
                                expense_entry["explanation"],
                                path,
                            ),
                        )
                        inserted_expense += 1

                cur.execute(
                    """
                    INSERT INTO files (path, hash, last_processed)
                    VALUES (?, ?, ?)
                    ON CONFLICT(path) DO UPDATE SET
                        hash=excluded.hash,
                        last_processed=excluded.last_processed
                    """,
                    (path, current_hash, datetime.utcnow().isoformat()),
                )

                conn.commit()

                processed += 1
                total_time_rows += inserted_time
                total_expense_rows += inserted_expense
                pbar.set_postfix({"time_rows": inserted_time, "expense_rows": inserted_expense})
                pbar.update(1)

            except Exception as e:
                conn.rollback()
                errors += 1
                pbar.set_postfix_str("error")
                pbar.update(1)
                tqdm.write(f"Error processing {path}: {e}")

    conn.close()

    tqdm.write("Rebase complete.")
    tqdm.write(f"Processed files: {processed}")
    tqdm.write(f"Skipped files: {skipped}")
    tqdm.write(f"Errors: {errors}")
    tqdm.write(f"Imported time rows: {total_time_rows}")
    tqdm.write(f"Imported expense rows: {total_expense_rows}")


# -----------------------------
# Search
# -----------------------------

def search(term: str):
    from colorama import Fore, Style, init

    init(autoreset=True)

    def highlight(text: str, search_term: str) -> str:
        if not search_term:
            return text
        pattern = re.compile(re.escape(search_term), re.IGNORECASE)
        return pattern.sub(lambda m: Fore.GREEN + m.group(0) + Style.RESET_ALL, text)

    conn = get_conn()
    cur = conn.cursor()
    like_term = f"%{term}%"

    print("--- Time Entries ---")
    for row in cur.execute(
        """
        SELECT *
        FROM time_entries
        WHERE CAST(date AS TEXT) LIKE ?
           OR job_number LIKE ?
           OR work_code LIKE ?
           OR description LIKE ?
           OR source_file LIKE ?
        """,
        (like_term, like_term, like_term, like_term, like_term),
    ):
        print(highlight(str(row), term))

    print("--- Expense Entries ---")
    for row in cur.execute(
        """
        SELECT
            e.id,
            e.time_entry_id,
            t.date,
            t.job_number,
            t.work_code,
            e.mileage,
            e.mileage_dollars,
            e.per_diem_food,
            e.air_fare,
            e.hotel,
            e.parking_tolls,
            e.rental_car_fuel,
            e.business_meals,
            e.other,
            e.explanation,
            e.source_file
        FROM expense_entries e
        JOIN time_entries t ON e.time_entry_id = t.id
        WHERE CAST(t.date AS TEXT) LIKE ?
           OR t.job_number LIKE ?
           OR t.work_code LIKE ?
           OR t.description LIKE ?
           OR CAST(e.mileage AS TEXT) LIKE ?
           OR CAST(e.mileage_dollars AS TEXT) LIKE ?
           OR CAST(e.per_diem_food AS TEXT) LIKE ?
           OR CAST(e.air_fare AS TEXT) LIKE ?
           OR CAST(e.hotel AS TEXT) LIKE ?
           OR CAST(e.parking_tolls AS TEXT) LIKE ?
           OR CAST(e.rental_car_fuel AS TEXT) LIKE ?
           OR CAST(e.business_meals AS TEXT) LIKE ?
           OR CAST(e.other AS TEXT) LIKE ?
           OR e.explanation LIKE ?
           OR e.source_file LIKE ?
        """,
        (
            like_term, like_term, like_term, like_term, like_term,
            like_term, like_term, like_term, like_term, like_term,
            like_term, like_term, like_term, like_term, like_term,
        ),
    ):
        print(highlight(str(row), term))

    conn.close()


# -----------------------------
# Query mode
# -----------------------------

def query_mode():
    """
    Interactive read-only SQL query mode.

    Only SELECT queries are allowed. This lets the user inspect time_entries,
    expense_entries, and joined data without modifying the database.
    """
    conn = get_conn()
    cur = conn.cursor()

    print("Query mode. SELECT statements only.")
    print("Available tables: time_entries, expense_entries")
    print("Type .tables to list tables, .schema to show schemas, .exit to quit.")

    while True:
        try:
            sql = input("query> ").strip()
        except (EOFError, KeyboardInterrupt):
            print()
            break

        if not sql:
            continue

        command = sql.lower()

        if command in (".exit", "exit", "quit", ".quit"):
            break

        if command == ".tables":
            for row in cur.execute(
                """
                SELECT name
                FROM sqlite_master
                WHERE type='table'
                ORDER BY name
                """
            ):
                print(row[0])
            continue

        if command == ".schema":
            for table_name in ("time_entries", "expense_entries"):
                print("-" * 80)
                print(table_name)
                for row in cur.execute(f"PRAGMA table_info({table_name})"):
                    print(row)
            continue

        if not command.startswith("select"):
            print("Only SELECT queries are allowed in query mode.")
            continue

        try:
            cur.execute(sql)
            rows = cur.fetchall()
            column_names = [desc[0] for desc in cur.description]

            if not rows:
                print("No rows returned.")
                continue

            print_table(column_names, rows)
            print(f"Rows returned: {len(rows)}")

        except sqlite3.Error as e:
            print(f"SQL error: {e}")

    conn.close()


# -----------------------------
# Reports
# -----------------------------

def last_report(offset: int = 0, show_time: bool = True, show_expense: bool = False):
    """Print a report for the recent imported timesheet file at the given offset."""
    conn = get_conn()
    cur = conn.cursor()

    if offset < 0:
        print("Offset must be 0 or greater.")
        conn.close()
        return

    cur.execute(
        """
        SELECT
            source_file,
            MAX(date) AS last_entry_date
        FROM time_entries
        WHERE date IS NOT NULL
          AND date != ''
        GROUP BY source_file
        ORDER BY last_entry_date DESC, source_file DESC
        LIMIT 1 OFFSET ?
        """,
        (offset,),
    )
    row = cur.fetchone()

    if not row:
        print(f"No imported timesheet found at offset {offset}.")
        conn.close()
        return

    source_file, last_entry_date = row

    cur.execute(
        """
        SELECT
            SUM(regular_hours),
            SUM(ot_hours),
            SUM(doubletime_hours)
        FROM time_entries
        WHERE source_file=?
        """,
        (source_file,),
    )
    hours = cur.fetchone()

    cur.execute(
        f"""
        SELECT SUM({expense_total_sql('e')})
        FROM expense_entries e
        JOIN time_entries t ON e.time_entry_id = t.id
        WHERE t.source_file=?
        """,
        (source_file,),
    )
    expenses = cur.fetchone()[0]

    regular_hours = hours[0] or 0
    ot_hours = hours[1] or 0
    doubletime_hours = hours[2] or 0
    total_hours = regular_hours + ot_hours + doubletime_hours

    print("Last Timesheet Report")
    print("=" * 80)
    print(f"Offset: {offset}")
    print(f"Source File: {source_file}")
    print(f"Last Entry Date: {last_entry_date}")
    print(f"Regular Hours: {regular_hours}")
    print(f"OT Hours: {ot_hours}")
    print(f"Doubletime Hours: {doubletime_hours}")
    print(f"Total Hours: {total_hours}")
    print(f"Total Expenses: {expenses or 0}")

    if show_time:
        print()
        print("Time Entries")
        print("=" * 80)
        cur.execute(
            """
            SELECT
                date,
                job_number,
                work_code,
                regular_hours,
                ot_hours,
                doubletime_hours,
                description,
                CASE overnight WHEN 1 THEN 'Yes' ELSE 'No' END AS overnight
            FROM time_entries
            WHERE source_file=?
            ORDER BY date ASC, id ASC
            """,
            (source_file,),
        )
        print_table(
            ["Date", "Job", "Code", "Reg", "OT", "DT", "Description", "Overnight"],
            cur.fetchall(),
        )

    if show_expense:
        print()
        print("Expense Entries")
        print("=" * 80)
        cur.execute(
            f"""
            SELECT
                t.date,
                t.job_number,
                t.work_code,
                e.mileage,
                e.mileage_dollars,
                e.per_diem_food,
                e.air_fare,
                e.hotel,
                e.parking_tolls,
                e.rental_car_fuel,
                e.business_meals,
                e.other,
                ({expense_total_sql('e')}) AS total,
                e.explanation
            FROM time_entries t
            LEFT JOIN expense_entries e ON e.time_entry_id = t.id
            WHERE t.source_file=?
            ORDER BY t.date ASC, t.id ASC
            """,
            (source_file,),
        )
        print_table(
            [
                "Date", "Job", "Code", "Miles", "Mileage $", "Per Diem",
                "Air Fare", "Hotel", "Parking/Tolls", "Rental/Fuel", "Meals",
                "Other", "Total", "Explanation",
            ],
            cur.fetchall(),
        )

    conn.close()


def project_report(job_number: str, verbose: bool = False):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute(
        """
        SELECT
            SUM(regular_hours),
            SUM(ot_hours),
            SUM(doubletime_hours)
        FROM time_entries
        WHERE job_number=?
        """,
        (job_number,),
    )
    hours = cur.fetchone()

    cur.execute(
        f"""
        SELECT SUM({expense_total_sql('e')})
        FROM expense_entries e
        JOIN time_entries t ON e.time_entry_id = t.id
        WHERE t.job_number=?
        """,
        (job_number,),
    )
    expenses = cur.fetchone()[0]

    regular_hours = hours[0] or 0
    ot_hours = hours[1] or 0
    doubletime_hours = hours[2] or 0
    total_hours = regular_hours + ot_hours + doubletime_hours

    print(f"Project: {job_number}")
    print(f"Regular Hours: {regular_hours}")
    print(f"OT Hours: {ot_hours}")
    print(f"Doubletime Hours: {doubletime_hours}")
    print(f"Total Hours: {total_hours}")
    print(f"Total Expenses: {expenses or 0}")

    if verbose:
        print()
        print("--- Verbose Project Entries (Grouped by File) ---")
        current_file = None

        cur.execute(
            f"""
            SELECT
                t.date,
                t.job_number,
                t.work_code,
                t.regular_hours,
                t.ot_hours,
                t.doubletime_hours,
                t.description,
                t.overnight,
                e.mileage,
                e.mileage_dollars,
                e.per_diem_food,
                e.air_fare,
                e.hotel,
                e.parking_tolls,
                e.rental_car_fuel,
                e.business_meals,
                e.other,
                ({expense_total_sql('e')}) AS expense_total,
                e.explanation,
                t.source_file
            FROM time_entries t
            LEFT JOIN expense_entries e ON e.time_entry_id = t.id
            WHERE t.job_number=?
            ORDER BY t.date ASC, t.source_file, t.id
            """,
            (job_number,),
        )
        rows = cur.fetchall()

        if not rows:
            print("No entries found for this project.")

        for row in rows:
            (
                entry_date,
                entry_job_number,
                work_code,
                regular,
                ot,
                doubletime,
                time_description,
                overnight,
                mileage,
                mileage_dollars,
                per_diem_food,
                air_fare,
                hotel,
                parking_tolls,
                rental_car_fuel,
                business_meals,
                other,
                expense_total,
                explanation,
                source_file,
            ) = row

            if source_file != current_file:
                current_file = source_file
                print("=" * 80)
                print(f"Source File: {source_file}")
                print("=" * 80)

            print("-" * 80)
            print(f"Date: {entry_date}")
            print(f"Job Number: {entry_job_number}")
            print(f"Work Code: {work_code}")
            print(f"Regular Hours: {regular}")
            print(f"OT Hours: {ot}")
            print(f"Doubletime Hours: {doubletime}")
            print(f"Time Description: {time_description}")
            print(f"Overnight: {'Yes' if overnight else 'No'}")
            print(f"Mileage: {mileage or 0}")
            print(f"Mileage Dollars: {mileage_dollars or 0}")
            print(f"Per Diem Food: {per_diem_food or 0}")
            print(f"Air Fare: {air_fare or 0}")
            print(f"Hotel: {hotel or 0}")
            print(f"Parking / Tolls: {parking_tolls or 0}")
            print(f"Rental Car / Fuel: {rental_car_fuel or 0}")
            print(f"Business Meals: {business_meals or 0}")
            print(f"Other: {other or 0}")
            print(f"Expense Total: {expense_total or 0}")
            print(f"Explanation: {explanation or ''}")

    conn.close()


# -----------------------------
# CLI
# -----------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Import, search, and report on timesheet / expense sheet data."
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"timesheet_expense_tool {__version__}",
    )

    sub = parser.add_subparsers(dest="command")

    cfg = sub.add_parser("config", help="Configure the timesheet directory")
    cfg.add_argument("directory")
    cfg.add_argument("--user", help="User name for timesheets", required=False)

    rebase_parser = sub.add_parser("rebase", help="Scan configured directory and import changed .xlsx files")
    rebase_parser.add_argument("-full", "--full", action="store_true", help="Clear database and fully rebuild")

    sub.add_parser("query", help="Enter interactive read-only SQL query mode")

    s = sub.add_parser("search", help="Search time and expense records")
    s.add_argument("term")

    last_parser = sub.add_parser("last", help="Show a report for a recent imported timesheet")
    last_parser.add_argument(
        "offset",
        nargs="?",
        type=int,
        default=0,
        help="How many sheets back to show. 0 = latest, 1 = previous, etc.",
    )
    last_parser.add_argument("-t", action="store_true", help="Show time entries (default)")
    last_parser.add_argument("-e", action="store_true", help="Show expense entries")

    pr = sub.add_parser("project_report", help="Show project totals")
    pr.add_argument("job_number")
    pr.add_argument("--verbose", action="store_true")

    args = parser.parse_args()

    init_db()

    if args.command == "config":
        set_config(args.directory, args.user)
    elif args.command == "rebase":
        if not validate_config():
            return
        if args.full:
            print("Performing full rebase: clearing existing imported data...")
            clear_imported_data()
        rebase()
    elif args.command == "search":
        if not validate_config():
            return
        search(args.term)
        search(args.term)
    elif args.command == "query":
        if not validate_config():
            return
        query_mode()
        query_mode()
    elif args.command == "last":
        if not validate_config():
            return
        show_time = args.t or not args.e
        show_expense = args.e
        last_report(offset=args.offset, show_time=show_time, show_expense=show_expense)
        show_time = args.t or not args.e
        show_expense = args.e
        last_report(offset=args.offset, show_time=show_time, show_expense=show_expense)
    elif args.command == "project_report":
        if not validate_config():
            return
        project_report(args.job_number, args.verbose)
        project_report(args.job_number, args.verbose)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
